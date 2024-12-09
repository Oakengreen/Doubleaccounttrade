
import openpyxl
import json

def excel_to_dictionary_with_formulas(file_path):
    # Ladda arbetsboken två gånger: en med beräknade värden och en med formler
    workbook_data_only = openpyxl.load_workbook(file_path, data_only=True)
    workbook_with_formulas = openpyxl.load_workbook(file_path, data_only=False)

    excel_dict = {}

    for sheet_name in workbook_data_only.sheetnames:
        sheet_data_only = workbook_data_only[sheet_name]
        sheet_with_formulas = workbook_with_formulas[sheet_name]

        sheet_data = {}
        for row in sheet_data_only.iter_rows():
            for cell in row:
                if cell.value is not None:  # Endast celler med data eller formler
                    formula_cell = sheet_with_formulas[cell.coordinate]
                    sheet_data[cell.coordinate] = {
                        "value": cell.value,  # Det beräknade värdet
                        "formula": formula_cell.value if formula_cell.data_type == 'f' else None,  # Formeln om den finns
                        "coordinates": cell.coordinate
                    }
        excel_dict[sheet_name] = sheet_data

    return excel_dict

# Använd funktionen
file_path = "DA (1).xlsx"
result = excel_to_dictionary_with_formulas(file_path)

# Skriv ut eller analysera dictionaryn
import pprint
pprint.pprint(result)

# Anta att result är den dictionary du genererat
with open("excel_data.json", "w") as json_file:
    json.dump(result, json_file, indent=4)

