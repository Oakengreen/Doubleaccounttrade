from openpyxl import load_workbook
import re
import networkx as nx
import matplotlib.pyplot as plt

# Filnamn till ditt Excel-ark
filename = "DA (1).xlsx"

# Öppna arbetsboken med formlerna intakta
wb = load_workbook(filename, data_only=False)

# Skapa en tom riktad graf
G = nx.DiGraph()

# Regex för att hitta cellreferenser i formler:
# ex: A1, B2, AA10 osv.
cell_ref_pattern = re.compile(r'([A-Z]{1,3}\d+)')

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            cell_ref = f"{sheet_name}!{cell.coordinate}"
            # Lägg till noden i grafen (om inte redan finns)
            G.add_node(cell_ref)

            # Kolla om cellen är en formel
            if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                # Hitta alla cellreferenser i formeln
                refs = cell_ref_pattern.findall(formula)

                # Lägg till kanter från refererade celler -> denna cell
                for r in refs:
                    # Vi antar här att referensen är i samma blad om inget annat anges
                    source = r
                    G.add_node(source)
                    G.add_edge(source, cell_ref)

# Nu har vi en beroendegraf G
# Rita grafen med networkx och matplotlib
plt.figure(figsize=(12, 8))
pos = nx.spring_layout(G, k=0.5, iterations=50)  # algoritm för placering av noderna
nx.draw(G, pos, with_labels=True, node_size=2000, font_size=8,
        node_color='lightblue', arrows=True, arrowstyle='-|>')
plt.title("Beroendegraf över Excel-formler")
plt.show()
